import plotly.express as px
import plotly.graph_objects as go
import akshare as ak
import os
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import multiprocessing
import datetime
import pandas as pd
import xlwings as xw


today = datetime.date.today().strftime('%Y%m%d')
os.chdir('/Users/liujiannan/Desktop/股票/ETF')
file_path = "sina_etf_list.xlsx"

def etf_show(etf, title):
    fig = px.line(etf, x="date", y="close", title=title)

    fig.add_trace(go.Scatter(x=[etf['date'].iloc[-1]],
                            y=[etf['close'].iloc[-1]],
                            text=[etf['date'].iloc[-1]],
                            mode='markers+text',
                            marker=dict(color='red', size=10),
                            textfont=dict(color='green', size=10),
                            textposition='top left',
                            showlegend=False))

    fig.add_trace(go.Scatter(x=[etf['date'].iloc[-1]],
                            y=[etf['close'].iloc[-1]],
                            text=[etf['close'].iloc[-1]],
                            mode='markers+text',
                            marker=dict(color='red', size=10),
                            textfont=dict(color='green', size=10),
                            textposition='bottom center',
                            showlegend=False))

    fig.show()


def ex_pic(file):
    wb = load_workbook(file)
    ws = wb.active
    columns = ws.max_column
    column_letter = get_column_letter((columns + 1))

    for i in range(2, len(lst) + 2):
        ws.column_dimensions[column_letter].width = 65
        ws.row_dimensions[i].height = 120
        if ws.cell(row=i, column=3).value:
            img = Image(str(ws.cell(row=i, column=3).value) + '.png')  # 缩放图片
            img.width, img.height = (500, 154)
            ws.add_image(img, (column_letter + str(i)))

    wb.save(file)  # 新的结果保存输出



if __name__ == "__main__":
    etf = ak.fund_etf_category_sina(symbol="ETF基金")
    # etf.to_excel("sina_etf_list.xlsx")

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        etf.to_excel(writer, today, index=False)

    xw.Book(file_path)
    # hkmi = ak.fund_etf_hist_sina(symbol="sh513060")
    # etf_show(hkmi)
    messagebox.showinfo('ETG', '...完成任务...')
