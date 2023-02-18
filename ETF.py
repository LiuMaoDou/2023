import tushare as ts
import pandas as pd
import time
import matplotlib.pyplot as plt
import os
import multiprocessing
import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

os.chdir('/Users/liujiannan/Desktop/ETF')
pro = ts.pro_api('8b334e2d5d482292aa8b516c7f59dd6d23c0266dac3dc407dfb63ea1')
day1 = datetime.date(2023, 2, 20).strftime('%Y%m%d')
if datetime.date.today().isoweekday() > 5:
    today = datetime.date.today() - datetime.timedelta(days=2)
    today = today.strftime('%Y%m%d')
else:
    today = datetime.date.today().strftime('%Y%m%d')


def tedliu(ts_code_p):
    # df = pro.fund_daily(ts_code=ts_code_p, start_date=day1, end_date=today, fields="ts_code,trade_date,pct_chg")
    df = pro.fund_daily(ts_code=ts_code_p, start_date=day1, end_date=today, fields="ts_code,trade_date,close")
    df.replace("2023", "", regex=True, inplace=True)
    df = df.T
    df.columns = df.iloc[1]
    df.drop("trade_date", inplace=True)
    # df.rename(index={"pct_chg": ts_code_p}, inplace=True)
    df.rename(index={"close": ts_code_p}, inplace=True)
    df.drop("ts_code", inplace=True)
    a = df.columns.values.tolist()
    b = df.loc[ts_code_p]
    df.loc['c'] = 0
    c = df.loc["c"]
    # plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False
    plt.xlabel('Daily')
    plt.ylabel('Value')
    plt.plot(a, b)
    plt.plot(a, c)
    plt.savefig((str(ts_code_p) + '.png'), dpi=100)
    plt.close()
    df.drop(['c'], inplace=True)
    return df


if __name__ == "__main__":
    print("...Starting...")
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    print(now)
    now1 = time.time()
    etf_data = pro.fund_daily(trade_date=today, fields="ts_code")
    pool = multiprocessing.Pool(processes=2)
    multiple_results = [pool.apply_async(tedliu, args=(i,)) for i in etf_data['ts_code']]
    lst = [res.get() for res in multiple_results]

    pool.close()
    pool.join()

    name = pro.fund_basic(market='E', fields="ts_code,name")
    amount = pro.fund_daily(trade_date=today, fields="ts_code,amount")

    result = pd.concat(lst)
    result.reset_index(inplace=True)
    result.rename(columns={'index': 'ts_code'}, inplace=True)
    result = pd.merge(result, name, on=["ts_code"])
    #
    ts_name = result['name']
    result.drop(labels=['name'], axis=1, inplace=True)
    result.insert(0, 'name', ts_name)

    result = pd.merge(result, amount, on=["ts_code"])
    result.sort_values("amount", ascending=False, inplace=True)

    result.to_excel('etf.xlsx')

    wb = load_workbook('etf.xlsx')
    ws = wb.active
    columns = ws.max_column
    column_letter = get_column_letter((columns + 1))

    for i in range(2, len(lst) + 2):
        ws.column_dimensions[column_letter].width = 65
        ws.row_dimensions[i].height = 120
        if ws.cell(row=i, column=3).value:
            img = Image(str(ws.cell(row=i, column=3).value) + '.png')  # 缩放图片
            img.width, img.height = (500, 154)
            ws.add_image(img, (column_letter + str(i)))

    wb.save('etf.xlsx')  # 新的结果保存输出

    print("------------------------------------")
    print("...Finished...")
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    print(now)
    now2 = time.time()
    now3 = now2 - now1
    print("...用时{} Secs...".format(round(now3, 2)))
